"""Machine validation for the official Q4 schedule and result workbook."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

import photography_model as photo
import shooting_model as shoot
from feasible_windows import _continuous_metrics, load_targets
from run_q4 import _untouched_snapshot
from trajectory_state import TrajectoryState


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    parser = argparse.ArgumentParser()
    parser.add_argument("--trajectory", type=Path, required=True)
    parser.add_argument("--targets", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--results", type=Path, required=True)
    parser.add_argument("--figures", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    params = json.loads((args.results / "parameters.json").read_text(encoding="utf-8"))
    schedule = pd.read_csv(args.results / "optimized_schedule.csv")
    trajectory = TrajectoryState.load(args.trajectory)
    targets = load_targets(args.targets)
    target_map = {target.target_id: target for target in targets}
    checks = {
        "selected_count_matches_optimum": len(schedule) == int(params["maximum_task_count"]),
        "fixed_nine_task_cap_removed": len(schedule) > 9,
        "time_sorted": bool(np.all(np.diff(schedule["任务执行时刻(s)"]) > 0)),
        "uncapped_model": params["milp"]["capacity_upper_bound"] is None,
        "all_milp_stages_optimal": all(
            int(params["milp"][key]) == 0
            for key in ("stage1_status", "stage2_status", "stage3_status")
        ),
        "all_milp_gaps_zero": all(
            abs(float(params["milp"][key])) <= 1e-12
            for key in ("stage1_gap", "stage2_gap", "stage3_gap")
        ),
        "positive_margin": bool(np.all(schedule["归一化最小裕度"] > 0)),
        "exact_count_beats_greedy": int(params["maximum_task_count"]) >= int(params["greedy_task_count"]),
    }
    intervals_ok = True
    for first in range(len(schedule)):
        for second in range(first + 1, len(schedule)):
            a, b = schedule.iloc[first], schedule.iloc[second]
            separated = (
                a["任务执行时刻(s)"] + 0.01 <= b["开始准备时刻(s)"] + 1e-9
                or b["任务执行时刻(s)"] + 0.01 <= a["开始准备时刻(s)"] + 1e-9
            )
            intervals_ok &= bool(separated)
    checks["preparation_intervals_disjoint"] = intervals_ok
    shooting = schedule[schedule["任务"] == shoot.TASK_NAME]
    checks["shooting_targets_unique"] = shooting["目标编号"].is_unique
    photo_angles_ok = True
    photography = schedule[schedule["任务"] == photo.TASK_NAME]
    for _, group in photography.groupby("目标编号"):
        values = group["方向角(deg)"].to_numpy()
        for i in range(len(values)):
            for j in range(i + 1, len(values)):
                photo_angles_ok &= photo.circular_separation_deg(values[i], values[j]) >= 60 - 1e-9
    checks["photography_angle_separation"] = photo_angles_ok

    continuous_ok = True
    reported_ok = True
    for _, row in schedule.iterrows():
        metrics = _continuous_metrics(
            trajectory, target_map[row["目标编号"]],
            float(row["开始准备时刻(s)"]),
            float(row["任务执行时刻(s)"]), 0.01,
        )
        continuous_ok &= metrics["margin"] >= -1e-9
        reported_ok &= abs(metrics["margin"] - float(row["归一化最小裕度"])) < 1e-10
    checks["continuous_001s_constraints"] = continuous_ok
    checks["reported_metrics_recomputed"] = reported_ok

    result_path = args.results / "result.xlsx"
    checks["untouched_template_semantics"] = (
        _untouched_snapshot(args.template) == _untouched_snapshot(result_path)
    )
    wb = load_workbook(result_path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    workbook_rows = [[ws.cell(row, column).value for column in range(1, 6)]
                     for row in range(2, len(schedule) + 2)]
    expected_rows = [[int(row["序号"]), row["目标编号"], row["任务"],
                      float(row["开始准备时刻(s)"]),
                      float(row["任务执行时刻(s)"])]
                     for _, row in schedule.iterrows()]
    checks["result_workbook_rows_match"] = workbook_rows == expected_rows
    checks["no_formula_errors"] = all(
        not (isinstance(cell.value, str) and cell.value.startswith("#"))
        for row in ws.iter_rows() for cell in row
    )
    stems = ["trajectory_targets_schedule", "candidate_feasibility_map",
             "optimized_schedule_timeline", "constraint_margins",
             "optimization_framework"]
    checks["figure_triplets_complete"] = all(
        (args.figures / f"{stem}.{suffix}").exists()
        for stem in stems for suffix in ("png", "svg", "pdf")
    )
    report = {"ok": bool(all(checks.values())), "checks": checks}
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    if not report["ok"]:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
