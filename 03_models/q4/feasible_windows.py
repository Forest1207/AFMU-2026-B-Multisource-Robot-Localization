"""Generate and continuously verify compressed Q4 task candidates."""

from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook

import photography_model as photo
import shooting_model as shoot
from trajectory_state import TrajectoryState


@dataclass(frozen=True)
class Target:
    target_id: str
    x_m: float
    y_m: float
    task: str


@dataclass(frozen=True)
class Candidate:
    candidate_id: int
    target_id: str
    task: str
    preparation_start_s: float
    execution_time_s: float
    angle_deg: float
    normalized_margin: float
    min_distance_m: float
    max_distance_m: float
    max_speed_mps: float
    max_acceleration_mps2: float

    def to_dict(self) -> dict:
        return asdict(self)


def load_targets(workbook: str | Path) -> list[Target]:
    wb = load_workbook(workbook, read_only=True, data_only=True)
    targets: list[Target] = []
    for sheet, task in (("射击目标", shoot.TASK_NAME), ("拍照目标", photo.TASK_NAME)):
        if sheet not in wb.sheetnames:
            raise ValueError(f"Missing target sheet: {sheet}")
        ws = wb[sheet]
        rows = []
        for row in range(2, ws.max_row + 1):
            values = [ws.cell(row, column).value for column in range(1, 4)]
            if all(value is None for value in values):
                continue
            if any(value is None for value in values):
                raise ValueError(f"Partially empty target row {sheet}!{row}")
            rows.append(Target(str(values[0]), float(values[1]), float(values[2]), task))
        if len(rows) != 18:
            raise ValueError(f"Expected 18 non-empty targets in {sheet}, found {len(rows)}")
        targets.extend(rows)
    if len({target.target_id for target in targets}) != 36:
        raise ValueError("Target identifiers must be unique.")
    return targets


def _task_config(task: str):
    if task == shoot.TASK_NAME:
        return shoot.PREPARATION_S, shoot.normalized_margin
    if task == photo.TASK_NAME:
        return photo.PREPARATION_S, photo.normalized_margin
    raise ValueError(f"Unknown task: {task}")


def _segments(indices: np.ndarray) -> list[np.ndarray]:
    if indices.size == 0:
        return []
    split = np.where(np.diff(indices) > 1)[0] + 1
    return [part for part in np.split(indices, split) if part.size]


def _compress_indices(indices: np.ndarray, margin: np.ndarray, angle: np.ndarray,
                      task: str, photo_angle_bin_deg: float = 5.0) -> np.ndarray:
    """Reference-style candidate compression.

    Shooting keeps each feasible segment's first, last and highest-margin time.
    Photography keeps the best candidate in every 5-degree bearing bin plus
    segment endpoints.  The 5-degree bin is deliberately much finer than the
    60-degree photo-separation constraint so compression does not erase useful
    angular alternatives near a bin boundary.
    """
    chosen: set[int] = set()
    segments = _segments(indices)
    if task == shoot.TASK_NAME:
        for segment in segments:
            chosen.add(int(segment[0]))
            chosen.add(int(segment[-1]))
            chosen.add(int(segment[np.argmax(margin[segment])]))
    elif task == photo.TASK_NAME:
        bins = np.floor((angle[indices] % 360.0) / photo_angle_bin_deg).astype(int)
        for angle_bin in np.unique(bins):
            members = indices[bins == angle_bin]
            chosen.add(int(members[np.argmax(margin[members])]))
        for segment in segments:
            chosen.add(int(segment[0]))
            chosen.add(int(segment[-1]))
    else:
        raise ValueError(f"Unknown task: {task}")
    return np.array(sorted(chosen), dtype=int)


def _continuous_metrics(trajectory: TrajectoryState, target: Target,
                        start: float, end: float, step: float = 0.01) -> dict:
    count = int(round((end - start) / step))
    query = start + step * np.arange(count + 1, dtype=float)
    query[-1] = end
    state = trajectory.evaluate(query)
    distance = np.hypot(state["x"] - target.x_m, state["y"] - target.y_m)
    _, margin_function = _task_config(target.task)
    margin = margin_function(distance, state["speed"], state["acceleration"])
    end_state = trajectory.evaluate(np.array([end]))
    angle = float(np.degrees(np.arctan2(
        end_state["y"][0] - target.y_m,
        end_state["x"][0] - target.x_m,
    )) % 360.0)
    return {
        "margin": float(np.min(margin)),
        "angle_deg": angle,
        "min_distance_m": float(np.min(distance)),
        "max_distance_m": float(np.max(distance)),
        "max_speed_mps": float(np.max(state["speed"])),
        "max_acceleration_mps2": float(np.max(state["acceleration"])),
    }


def generate_candidates(trajectory: TrajectoryState, targets: list[Target],
                        fine_step_s: float = 0.01,
                        photo_angle_bin_deg: float = 5.0) -> list[Candidate]:
    frame = trajectory.frame
    time = trajectory.time
    candidates: list[Candidate] = []
    next_id = 0
    for target in targets:
        preparation, margin_function = _task_config(target.task)
        distance = np.hypot(frame["x"] - target.x_m, frame["y"] - target.y_m).to_numpy()
        point_margin = margin_function(
            distance,
            frame["speed"].to_numpy(dtype=float),
            frame["acceleration"].to_numpy(dtype=float),
        )
        angle = np.degrees(np.arctan2(frame["y"] - target.y_m,
                                      frame["x"] - target.x_m)).to_numpy() % 360.0
        window_points = int(round(preparation / 0.1)) + 1
        rolling_min = np.full(time.size, -np.inf)
        for end_index in range(window_points - 1, time.size):
            start_index = end_index - window_points + 1
            rolling_min[end_index] = float(np.min(point_margin[start_index:end_index + 1]))
        feasible = np.where(rolling_min >= -1e-12)[0]
        compressed = _compress_indices(
            feasible, rolling_min, angle, target.task,
            photo_angle_bin_deg=photo_angle_bin_deg,
        )
        for end_index in compressed:
            end = float(time[end_index])
            start = end - preparation
            if start < time[0] - 1e-9:
                continue
            metrics = _continuous_metrics(trajectory, target, start, end, fine_step_s)
            if metrics["margin"] < -1e-9:
                continue
            candidates.append(Candidate(
                candidate_id=next_id,
                target_id=target.target_id,
                task=target.task,
                preparation_start_s=start,
                execution_time_s=end,
                angle_deg=metrics["angle_deg"],
                normalized_margin=metrics["margin"],
                min_distance_m=metrics["min_distance_m"],
                max_distance_m=metrics["max_distance_m"],
                max_speed_mps=metrics["max_speed_mps"],
                max_acceleration_mps2=metrics["max_acceleration_mps2"],
            ))
            next_id += 1
    return candidates
