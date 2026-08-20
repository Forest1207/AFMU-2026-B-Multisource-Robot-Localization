"""Run Q4 candidate generation, lexicographic MILP and result export.

Formal Q4 follows the problem statement/reference-package interpretation:
there is no nine-task capacity and no cross-task preparation-time mutex.  The
Excel template is extended downward when the optimal plan contains more rows;
its red instruction/example region is preserved byte-for-semantics at the cell
value/style level.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sys
from copy import copy
from dataclasses import replace
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.interpolate import PchipInterpolator

import photography_model as photo
import shooting_model as shoot
from feasible_windows import Candidate, _continuous_metrics, generate_candidates, load_targets
from scheduler import optimize_schedule
from trajectory_state import TrajectoryState


def sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _cell_signature(cell) -> dict:
    return {
        "value": cell.value,
        "style_id": cell.style_id,
        "number_format": cell.number_format,
        "font": str(cell.font),
        "fill": str(cell.fill),
        "border": str(cell.border),
        "alignment": str(cell.alignment),
        "protection": str(cell.protection),
    }


def _protected_snapshot(path: str | Path) -> dict:
    """Snapshot cells that the submission writer is never allowed to change.

    A:E below row 1 are the expandable result table.  The original header and
    all H:L instruction/example cells are protected.  This matches the problem
    statement's explicit requirement not to change the red text without
    inventing a nine-row capacity that is not stated in the task.
    """
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    cells = {}
    for column in range(1, 6):
        cell = ws.cell(1, column)
        cells[cell.coordinate] = _cell_signature(cell)
    for row in range(1, ws.max_row + 1):
        for column in range(8, min(ws.max_column, 12) + 1):
            cell = ws.cell(row, column)
            if cell.value is not None or cell.style_id != 0:
                cells[cell.coordinate] = _cell_signature(cell)
    return {
        "sheetnames": wb.sheetnames,
        "merged": [str(item) for item in ws.merged_cells.ranges],
        "freeze_panes": str(ws.freeze_panes),
        "cells": cells,
    }


# Backward-compatible name used by older validation tooling.
_untouched_snapshot = _protected_snapshot


def _rounded_and_verified(selected: list[Candidate], trajectory: TrajectoryState,
                          targets) -> list[Candidate]:
    target_map = {target.target_id: target for target in targets}
    rounded = []
    for candidate in selected:
        end = round(candidate.execution_time_s, 2)
        preparation = shoot.PREPARATION_S if candidate.task == shoot.TASK_NAME else photo.PREPARATION_S
        start = round(end - preparation, 2)
        metrics = _continuous_metrics(
            trajectory, target_map[candidate.target_id], start, end, 0.01
        )
        if metrics["margin"] < -1e-9:
            raise RuntimeError(
                f"Two-decimal schedule time failed 0.01 s recheck: "
                f"{candidate.task}/{candidate.target_id}@{end:.2f}"
            )
        rounded.append(replace(
            candidate,
            preparation_start_s=start,
            execution_time_s=end,
            angle_deg=metrics["angle_deg"],
            normalized_margin=metrics["margin"],
            min_distance_m=metrics["min_distance_m"],
            max_distance_m=metrics["max_distance_m"],
            max_speed_mps=metrics["max_speed_mps"],
            max_acceleration_mps2=metrics["max_acceleration_mps2"],
        ))
    rounded.sort(key=lambda item: (item.execution_time_s, item.task, item.target_id))
    return rounded


def greedy_baseline(candidates: list[Candidate]) -> list[Candidate]:
    """Reference-style baseline: one highest-margin candidate per feasible target.

    For photography this deliberately keeps only one photo per target; the MILP
    can then demonstrate the value of its second lexicographic objective by
    selecting additional views at >=60 degree separation without sacrificing
    first-level target coverage.
    """
    groups: dict[tuple[str, str], list[Candidate]] = {}
    for candidate in candidates:
        groups.setdefault((candidate.task, candidate.target_id), []).append(candidate)
    selected = [max(group, key=lambda item: item.normalized_margin) for group in groups.values()]
    selected.sort(key=lambda item: (item.execution_time_s, item.task, item.target_id))
    return selected


def robustness_check(selected: list[Candidate], trajectory: TrajectoryState,
                     targets, trials: int = 500, seed: int = 2026,
                     position_std_scale: float = 0.5,
                     state_relative_sd: float = 0.05) -> dict:
    rng = np.random.default_rng(seed)
    frame = trajectory.frame
    sx_interp = PchipInterpolator(trajectory.time, frame["x_std"], extrapolate=False)
    sy_interp = PchipInterpolator(trajectory.time, frame["y_std"], extrapolate=False)
    target_map = {target.target_id: target for target in targets}
    task_success = np.ones((trials, len(selected)), dtype=bool)
    for index, candidate in enumerate(selected):
        target = target_map[candidate.target_id]
        count = int(round((candidate.execution_time_s - candidate.preparation_start_s) / 0.01))
        query = candidate.preparation_start_s + 0.01 * np.arange(count + 1)
        query[-1] = candidate.execution_time_s
        state = trajectory.evaluate(query)
        sx = np.asarray(sx_interp(query), dtype=float)
        sy = np.asarray(sy_interp(query), dtype=float)
        for trial in range(trials):
            x = state["x"] + rng.normal(0.0, position_std_scale) * sx
            y = state["y"] + rng.normal(0.0, position_std_scale) * sy
            speed = state["speed"] * max(0.0, 1.0 + rng.normal(0.0, state_relative_sd))
            acceleration = state["acceleration"] * max(0.0, 1.0 + rng.normal(0.0, state_relative_sd))
            distance = np.hypot(x - target.x_m, y - target.y_m)
            margin_function = (
                shoot.normalized_margin if candidate.task == shoot.TASK_NAME
                else photo.normalized_margin
            )
            task_success[trial, index] = bool(np.min(
                margin_function(distance, speed, acceleration)
            ) >= 0.0)
    return {
        "trials": trials,
        "seed": seed,
        "position_std_scale": position_std_scale,
        "speed_acceleration_relative_sd": state_relative_sd,
        "per_task_feasible_rate": task_success.mean(axis=0).tolist(),
        "whole_schedule_feasible_rate": float(np.all(task_success, axis=1).mean()),
    }


def _copy_row_style(ws, source_row: int, target_row: int, max_column: int = 5) -> None:
    for column in range(1, max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
    if source_row in ws.row_dimensions:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height


def write_result_template(template: str | Path, output: str | Path,
                          selected: list[Candidate]) -> dict:
    before = _protected_snapshot(template)
    output = Path(output)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, output)
    wb = load_workbook(output)
    ws = wb[wb.sheetnames[0]]

    # Clear only the original editable result cells.  Rows are then extended as
    # needed; the red instruction/example area H:L is never touched.
    original_rows = ws.max_row
    for row in range(2, original_rows + 1):
        for column in range(1, 6):
            ws.cell(row, column).value = None

    style_source_row = min(max(2, original_rows), 10)
    for row, candidate in enumerate(selected, start=2):
        if row > original_rows:
            _copy_row_style(ws, style_source_row, row, max_column=5)
        ws.cell(row, 1).value = row - 1
        ws.cell(row, 2).value = candidate.target_id
        ws.cell(row, 3).value = candidate.task
        ws.cell(row, 4).value = candidate.preparation_start_s
        ws.cell(row, 5).value = candidate.execution_time_s
    wb.save(output)

    after = _protected_snapshot(output)
    if before != after:
        raise RuntimeError("Protected result.xlsx header or red instruction/example cells changed.")
    return {
        "template_sha256": sha256(template),
        "output_sha256": sha256(output),
        "protected_snapshot_equal": True,
        "writable_region": "A:E, rows >= 2 (expandable)",
        "filled_rows": len(selected),
        "template_rows_before_expansion": original_rows,
    }


def _write_summary(path: Path, parameters: dict) -> None:
    ratio = (
        parameters["selected_task_count"] / parameters["greedy_task_count"]
        if parameters["greedy_task_count"] else float("nan")
    )
    text = f"""# 问题四正式结果摘要

- 主模型不设置人为的 9 项任务容量，也不加入题面未给出的跨任务准备时间互斥约束。
- 一级目标：最大化可覆盖目标数；二级目标：覆盖数固定后最大化满足 60° 角差的有效拍照次数；三级目标：前两级固定后最大化总安全裕度。
- 连续复核候选数：**{parameters['candidate_count']}**；最终覆盖目标数：**{parameters['coverage_count']}**。
- 最终任务记录共 **{parameters['selected_task_count']}** 条，其中射击 **{parameters['shooting_count']}** 次、拍照 **{parameters['photography_count']}** 次；射击期望命中数为 **{parameters['expected_shooting_hits']:.2f}**。
- 贪心基线按每个可行目标独立选取一个最大安全裕度时刻，共 {parameters['greedy_task_count']} 条记录；MILP 在不牺牲一级覆盖目标数的前提下通过多角度拍照增加有效任务记录。
- 三阶段 HiGHS MILP 均要求 0 relative gap；最终两位小数时刻再次按 0.01 s 完整准备窗口复核。
- `result.xlsx` 允许在 A:E 向下扩展结果行；原模板表头和 H:L 红色说明/范例保持不变。模板中的初始 9 行不解释为题目任务容量。
"""
    path.write_text(text, encoding="utf-8")


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    parser = argparse.ArgumentParser()
    parser.add_argument("--trajectory", type=Path, required=True)
    parser.add_argument("--targets", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--results", type=Path, required=True)
    args = parser.parse_args()
    args.results.mkdir(parents=True, exist_ok=True)

    trajectory = TrajectoryState.load(args.trajectory)
    targets = load_targets(args.targets)
    candidates = generate_candidates(trajectory, targets)
    schedule = optimize_schedule(candidates)
    selected = _rounded_and_verified(schedule.selected, trajectory, targets)
    greedy = greedy_baseline(candidates)

    robust_scenarios = {
        "moderate": robustness_check(
            selected, trajectory, targets, trials=500, seed=2026,
            position_std_scale=0.5, state_relative_sd=0.05,
        ),
        "standard": robustness_check(
            selected, trajectory, targets, trials=500, seed=2027,
            position_std_scale=1.0, state_relative_sd=0.05,
        ),
        "severe": robustness_check(
            selected, trajectory, targets, trials=500, seed=2028,
            position_std_scale=1.5, state_relative_sd=0.10,
        ),
    }

    candidate_frame = pd.DataFrame([candidate.to_dict() for candidate in candidates])
    candidate_frame.to_csv(args.results / "feasible_tasks.csv", index=False, encoding="utf-8-sig")
    schedule_frame = pd.DataFrame([
        {
            "序号": index,
            "目标编号": item.target_id,
            "任务": item.task,
            "开始准备时刻(s)": item.preparation_start_s,
            "任务执行时刻(s)": item.execution_time_s,
            "方向角(deg)": item.angle_deg,
            "归一化最小裕度": item.normalized_margin,
            "准备窗最小距离(m)": item.min_distance_m,
            "准备窗最大距离(m)": item.max_distance_m,
            "准备窗最大速度(m/s)": item.max_speed_mps,
            "准备窗最大加速度(m/s²)": item.max_acceleration_mps2,
        }
        for index, item in enumerate(selected, start=1)
    ])
    schedule_frame.to_csv(args.results / "optimized_schedule.csv", index=False, encoding="utf-8-sig")

    workbook = write_result_template(args.template, args.results / "result.xlsx", selected)
    shot_count = sum(item.task == shoot.TASK_NAME for item in selected)
    photo_count = len(selected) - shot_count
    greedy_shot_count = sum(item.task == shoot.TASK_NAME for item in greedy)
    greedy_photo_count = len(greedy) - greedy_shot_count
    greedy_coverage = len({(item.task, item.target_id) for item in greedy})

    parameters = {
        "trajectory": str(args.trajectory.resolve()),
        "trajectory_sha256": sha256(args.trajectory),
        "targets_workbook": str(args.targets.resolve()),
        "targets_sha256": sha256(args.targets),
        "target_counts": {"shooting": 18, "photography": 18},
        "candidate_count": len(candidates),
        "candidate_count_by_task": candidate_frame.groupby("task").size().to_dict(),
        "coverage_count": schedule.coverage_count,
        "selected_task_count": len(selected),
        "shooting_count": shot_count,
        "photography_count": photo_count,
        "expected_shooting_hits": shoot.HIT_PROBABILITY * shot_count,
        "minimum_normalized_margin": min(item.normalized_margin for item in selected),
        "total_normalized_margin": sum(item.normalized_margin for item in selected),
        "greedy_coverage_count": greedy_coverage,
        "greedy_task_count": len(greedy),
        "greedy_shooting_count": greedy_shot_count,
        "greedy_photography_count": greedy_photo_count,
        "greedy_minimum_margin": min(item.normalized_margin for item in greedy),
        "milp": {
            "method": "lexicographic target coverage -> photo count -> total safety margin",
            "artificial_capacity": None,
            "cross_task_time_mutex": False,
            "angle_conflict_count": schedule.angle_conflict_count,
            "stage1_status": schedule.stage1_status,
            "stage2_status": schedule.stage2_status,
            "stage3_status": schedule.stage3_status,
            "stage1_gap": schedule.stage1_gap,
            "stage2_gap": schedule.stage2_gap,
            "stage3_gap": schedule.stage3_gap,
        },
        "continuous_recheck_step_s": 0.01,
        "robustness_scenarios": robust_scenarios,
        "result_workbook": workbook,
    }
    (args.results / "parameters.json").write_text(
        json.dumps(parameters, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    _write_summary(args.results / "summary.md", parameters)
    print(json.dumps(parameters, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
